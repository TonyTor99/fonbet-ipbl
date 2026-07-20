"""Выгрузка снимков сборщика шорт-хоккея в Excel.

Строка = матч × игровая минута (плюс строка «до матча»). По каждому рынку:
линия, кф каждого исхода и результат (В/П/Возврат) рядом с кф.

Запуск:
    python export_shorthockey.py                 # -> export_shorthockey_YYYYMMDD_HHMMSS.xlsx
    python export_shorthockey.py /path/file.xlsx
"""
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import sh_collector_db as db

# (Заголовок, ключ строки БД). Ключи с "__" — спец-обработка ниже.
COLUMNS = [
    ("Дата МСК", "__date"),
    ("Время МСК", "__time"),
    ("Лига", "league"),
    ("Команда 1", "team1"),
    ("Команда 2", "team2"),
    ("До матча", "__prematch"),
    ("Игр. мин.", "__minute"),
    ("Период", "period"),
    ("Периоды", "periods"),
    ("Счёт", "__score"),
    # Исход 1X2
    ("П1 кф", "win1_odds"), ("П1", "r_win1"),
    ("X кф", "draw_odds"),  ("X", "r_draw"),
    ("П2 кф", "win2_odds"), ("П2", "r_win2"),
    # Двойные шансы
    ("1X кф", "dc_1x_odds"), ("1X", "r_1x"),
    ("12 кф", "dc_12_odds"), ("12", "r_12"),
    ("X2 кф", "dc_x2_odds"), ("X2", "r_x2"),
    # Фора
    ("Фора К1", "fora_line"), ("Фора К2", "__fora2_line"),
    ("Ф1 кф", "fora1_odds"), ("Ф1", "r_fora1"),
    ("Ф2 кф", "fora2_odds"), ("Ф2", "r_fora2"),
    # Тотал
    ("Тотал", "total_line"),
    ("ТБ кф", "total_b_odds"), ("ТБ", "r_total_b"),
    ("ТМ кф", "total_m_odds"), ("ТМ", "r_total_m"),
    # Инд. тоталы
    ("ИТ1", "it1_line"),
    ("ИТБ1 кф", "it1_b_odds"), ("ИТБ1", "r_it1_b"),
    ("ИТМ1 кф", "it1_m_odds"), ("ИТМ1", "r_it1_m"),
    ("ИТ2", "it2_line"),
    ("ИТБ2 кф", "it2_b_odds"), ("ИТБ2", "r_it2_b"),
    ("ИТМ2 кф", "it2_m_odds"), ("ИТМ2", "r_it2_m"),
    # Итог
    ("Итог счёт", "final_score"),
    ("Итог тотал", "final_total"),
]

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")
WIN_FILL = PatternFill("solid", fgColor="C6EFCE")     # зелёный
LOSE_FILL = PatternFill("solid", fgColor="FFC7CE")    # красный
PUSH_FILL = PatternFill("solid", fgColor="FFEB9C")    # жёлтый (возврат)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RESULT_KEYS = {h for h, k in COLUMNS if k and k.startswith("r_")}


def _value(row: dict, key: str):
    if key == "__score":
        return f"{row['score1']}:{row['score2']}"
    if key == "__prematch":
        return "да" if row.get("is_prematch") else None
    if key == "__minute":
        return "Не начался" if row.get("is_prematch") else row.get("game_minute")
    if key == "__date":
        return (row.get("snap_dt_msk") or "").split(" ")[0] or None
    if key == "__time":
        parts = (row.get("snap_dt_msk") or "").split(" ")
        return parts[1] if len(parts) > 1 else None
    if key == "__fora2_line":
        fl = row.get("fora_line")
        return -fl if fl is not None else None
    return row.get(key)


def _fill_for(val: str):
    if val == "Выигрыш":
        return WIN_FILL
    if val == "Проигрыш":
        return LOSE_FILL
    if val == "Возврат":
        return PUSH_FILL
    return None


def build(path: str):
    rows = db.all_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "Рынки шорт-хоккей"

    for c, (head, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, head)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, 2):
        for c, (head, key) in enumerate(COLUMNS, 1):
            val = _value(row, key)
            cell = ws.cell(r, c, val)
            cell.border = BORDER
            if head in RESULT_KEYS and val:
                fill = _fill_for(val)
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(horizontal="center")

    for c, (head, _) in enumerate(COLUMNS, 1):
        letter = ws.cell(1, c).column_letter
        ws.column_dimensions[letter].width = max(8, min(20, len(head) + 2))

    wb.save(path)
    st = db.stats()
    print(f"Сохранено: {path}")
    print(f"Строк: {st['rows']} | матчей: {st['events']} | с результатом: {st['resolved']}")


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else \
        f"export_shorthockey_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    db.init_db()
    build(path)


if __name__ == "__main__":
    main()
