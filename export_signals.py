"""Выгрузка ТМ-сигналов (стратегия signal_tm) в Excel.

Строка = запись по матчу на перерыве. Колонки-флаги «Прошёл формулу» и
«Отправлен в ТГ» позволяют отделить реальные сигналы от просто снимков перерыва.
На таблицу включён автофильтр — фильтровать можно прямо в Excel.

Запуск:
    python export_signals.py                 # -> export_signals_YYYYMMDD_HHMMSS.xlsx
    python export_signals.py /path/file.xlsx
"""
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import database

STRATEGY = "signal_tm"

# (Заголовок, ключ строки БД либо спец-ключ "__..." -> обработка в _value)
COLUMNS = [
    ("Дата-время", "created_at"),
    ("Прошёл формулу", "__qualified"),
    ("Отправлен в ТГ", "__sent"),
    ("В окне работы", "__in_window"),
    ("Лига", "league"),
    ("Команда 1", "team1"),
    ("Команда 2", "team2"),
    ("Счёт перерыва", "__half_score"),
    ("Q1", "q1"),
    ("Q2", "q2"),
    ("Q3", "q3"),
    ("Q4", "q4"),
    ("Сумма к перерыву", "half_total"),
    ("ТМ прематч", "line_prematch"),
    ("ТМ сигнал", "line"),
    ("Кф", "odds"),
    ("Формула (2×сумма−линия)", "formula_value"),
    ("Итог счёт", "final_score"),
    ("Итог тотал", "final_total"),
    ("Результат", "result"),
    ("Прибыль ₽", "profit"),
    ("Блок ТМ (перерыв)", "totals_snapshot"),
]

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")
WIN_FILL = PatternFill("solid", fgColor="C6EFCE")    # зелёный
LOSE_FILL = PatternFill("solid", fgColor="FFC7CE")   # красный
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = {"__qualified", "__sent", "__in_window"}


def _value(row: dict, key: str):
    if key == "__qualified":
        return "Да" if row.get("qualified") else "Нет"
    if key == "__sent":
        return "Да" if row.get("status") == "sent" else "Нет"
    if key == "__in_window":
        return "Да" if row.get("in_window") else "Нет"
    if key == "__half_score":
        return f"{row.get('fixed_score1')}:{row.get('fixed_score2')}"
    return row.get(key)


def build(path: str) -> int:
    """Строит xlsx, возвращает число строк-записей."""
    rows = database.all_signals(STRATEGY)
    wb = Workbook()
    ws = wb.active
    ws.title = "Сигналы ТМ"

    # шапка
    for c, (head, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, head)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = "A2"

    # данные
    for r, row in enumerate(rows, 2):
        for c, (head, key) in enumerate(COLUMNS, 1):
            val = _value(row, key)
            cell = ws.cell(r, c, val)
            cell.border = BORDER
            if head == "Результат" and val:
                cell.fill = WIN_FILL if val == "Выигрыш" else LOSE_FILL
                cell.alignment = Alignment(horizontal="center")
            elif key == "__qualified":
                cell.fill = WIN_FILL if val == "Да" else LOSE_FILL
                cell.alignment = Alignment(horizontal="center")
            elif key in CENTER:
                cell.alignment = Alignment(horizontal="center")

    # ширины + автофильтр
    for c, (head, _) in enumerate(COLUMNS, 1):
        letter = ws.cell(1, c).column_letter
        ws.column_dimensions[letter].width = max(10, min(26, len(head) + 2))
    last = ws.cell(1, len(COLUMNS)).column_letter
    ws.auto_filter.ref = f"A1:{last}{len(rows) + 1}"

    wb.save(path)
    return len(rows)


def main():
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        path = f"export_signals_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    database.init_db()
    n = build(path)
    print(f"Сохранено: {path} | строк: {n}")


if __name__ == "__main__":
    main()
