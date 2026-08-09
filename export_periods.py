"""Выгрузка снимков сборщика рынков ПО ЧЕТВЕРТЯМ в Excel.

Строка = матч × игровая минута × четверть. По каждому рынку четверти: линия, кф
исходов и результат (В/П) рядом с кф. В четверти, в отличие от матча, есть
ничья (X). Результат считается по счёту КОНКРЕТНОЙ четверти (колонка «Счёт четв.»).

Запуск:
    python export_periods.py                          # -> export_periods_...xlsx (Pro муж)
    python export_periods.py /path/file.xlsx db.db
"""
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import collector_periods_db

# (Заголовок, ключ строки БД). None-ключ -> спец-обработка ниже.
COLUMNS = [
    ("Дата МСК", "__date"),
    ("Время МСК", "__time"),
    ("Лига", "league"),
    ("Команда 1", "team1"),
    ("Команда 2", "team2"),
    ("Игр. мин.", "game_minute"),
    ("Четверть", "quarter"),
    ("Счёт матча", "__score"),
    ("Счёт четв. тек.", "q_live_score"),
    ("Фора К1", "fora_line"),
    ("Фора К2", "__fora2_line"),
    ("Фора П1 кф", "fora1_odds"),
    ("Фора П1", "r_fora1"),
    ("Фора П2 кф", "fora2_odds"),
    ("Фора П2", "r_fora2"),
    ("Тотал", "total_line"),
    ("ТБ кф", "total_b_odds"),
    ("ТБ", "r_total_b"),
    ("ТМ кф", "total_m_odds"),
    ("ТМ", "r_total_m"),
    ("ИТ1", "it1_line"),
    ("ИТБ1 КФ", "it1_b_odds"),
    ("ИТБ1 Р", "r_it1_b"),
    ("ИТМ1 КФ", "it1_m_odds"),
    ("ИТМ1 Р", "r_it1_m"),
    ("ИТ2", "it2_line"),
    ("ИТБ2 КФ", "it2_b_odds"),
    ("ИТБ2 Р", "r_it2_b"),
    ("ИТМ2 КФ", "it2_m_odds"),
    ("ИТМ2 Р", "r_it2_m"),
    ("П1 кф", "win1_odds"),
    ("П1", "r_win1"),
    ("X кф", "winx_odds"),
    ("X", "r_winx"),
    ("П2 кф", "win2_odds"),
    ("П2", "r_win2"),
    ("Счёт четв.", "q_score"),
    ("Итог счёт", "final_score"),
    ("Итог тотал", "final_total"),
]

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")
WIN_FILL = PatternFill("solid", fgColor="C6EFCE")    # зелёный
LOSE_FILL = PatternFill("solid", fgColor="FFC7CE")   # красный
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RESULT_KEYS = {h for h, k in COLUMNS if k and k.startswith("r_")}


def _value(row: dict, key: str):
    if key == "__score":
        return f"{row['score1']}:{row['score2']}"
    if key == "__date":
        return (row.get("snap_dt_msk") or "").split(" ")[0] or None
    if key == "__time":
        parts = (row.get("snap_dt_msk") or "").split(" ")
        return parts[1] if len(parts) > 1 else None
    if key == "__fora2_line":
        fl = row.get("fora_line")
        return -fl if fl is not None else None
    return row.get(key)


def build(path: str, db: str, title: str = "Рынки четвертей IPBL") -> int:
    """Выгружает снимки лиги из БД `db` в Excel. Возвращает число строк."""
    rows = collector_periods_db.all_rows(db)
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]   # лимит имени листа Excel

    # шапка
    for c, (head, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, head)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = "A2"

    # данные
    for r, row in enumerate(rows, 2):
        for c, (head, key) in enumerate(COLUMNS, 1):
            val = _value(row, key)
            cell = ws.cell(r, c, val)
            cell.border = BORDER
            if head in RESULT_KEYS and val:
                cell.fill = WIN_FILL if val == "Выигрыш" else LOSE_FILL
                cell.alignment = Alignment(horizontal="center")

    # ширины
    for c, (head, _) in enumerate(COLUMNS, 1):
        letter = ws.cell(1, c).column_letter
        ws.column_dimensions[letter].width = max(9, min(20, len(head) + 2))

    wb.save(path)
    st = collector_periods_db.stats(db)
    print(f"Сохранено: {path}")
    print(f"Строк: {st['rows']} | матчей: {st['events']} | с результатом: {st['resolved']}")
    return st["rows"]


def main():
    # По умолчанию — Pro муж. Можно указать: export_periods.py [файл.xlsx] [db.db]
    db = sys.argv[2] if len(sys.argv) > 2 else "pro_periods.db"
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        path = f"export_periods_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    collector_periods_db.init_db(db)
    build(path, db, "Рынки четвертей IPBL")


if __name__ == "__main__":
    main()
